from bs4 import BeautifulSoup as bs
from pathlib import Path
from typing import Optional,Union,Dict,List
from openpyxl import Workbook
import time
import os
import re
import requests as rq
import json
import numpy as np



def get_headers(
    key: str,
    default_value: Optional[str] = None
    )-> Dict[str,Dict[str,str]]:
    """ Get Headers """
    JSON_FILE : str = 'json/headers.json'

    with open(JSON_FILE,'r',encoding='UTF-8') as file:
        headers : Dict[str,Dict[str,str]] = json.loads(file.read())

    try :
        return headers[key]
    except:
        if default_value:
            return default_value
        raise EnvironmentError(f'Set the {key}')

class Coupang:
    @staticmethod
    def get_product_code(url: str)-> str:
        """ 입력받은 URL 주소의 PRODUCT CODE 추출하는 메소드 """
        prod_code : str = url.split('products/')[-1].split('?')[0]
        return prod_code

    def __init__(self)-> None:
        self.__headers : Dict[str,str] = get_headers(key='headers')
        self.page = 1
        self.star5 = 0
        self.star3 = 0
        self.star1 = 0


    def main(self)-> List[List[Dict[str,Union[str,int]]]]:
        
        # global star5_cnt
        # star5_cnt = 0
        # URL 주소
        URL : str = self.input_review_url()

        # URL의 Product Code 추출
        prod_code : str = self.get_product_code(url=URL)

        ###############
        url = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page=1&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true']
        self.__headers['referer'] = URL
        with rq.Session() as session:
            with session.get(url=url[0],headers=self.__headers) as response :
                html = response.text
                soup = bs(html,'html.parser')
                test = soup.select("div.js_reviewArticleHiddenValue")
                total_review = 0
                for i in range(len(test)):
                    total_review += int(test[i].attrs['data-count'])
                total_page = int(np.ceil(total_review/5))
                print("(이야야야야ㅏㅇ) ", total_review, int(np.ceil(total_review/5)))
        #################

        # URL 주소 재가공
        URLS : List[str] = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true' for page in range(1,total_page + 1)]
        """
        while (self.star1 + self.star3 + self.star5 <= 150):
            print("page ", self.page)
            URLS : List[str] = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={self.page}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true']
            self.__headers['referer'] = URL
            self.page += 1
                
            if (self.star1 + self.star3 + self.star5 == 150):
                break
        """
        # __headers에 referer 키 추가
        self.__headers['referer'] = URL
        
        with rq.Session() as session:
            return [self.fetch(url=url,session=session) for url in URLS]

    def fetch(self,url:str,session)-> List[Dict[str,Union[str,int]]]:
        save_data : List[Dict[str,Union[str,int]]] = list()
        print("@함수 시작")
        # url = [f'https://www.coupang.com/vp/product/reviews?productId={prod_code}&page={1}&size=5&sortBy=ORDER_SCORE_ASC&ratings=&q=&viRoleCode=3&ratingSummary=true']
        

        with session.get(url=url,headers=self.__headers) as response :
            # print("@with 시작")
            html = response.text
            soup = bs(html,'html.parser')

            # 총 리뷰 수로 총 페이지 수 계산하기
            # test = soup.select("div.js_reviewArticleHiddenValue")
            # total_review = 0
            # for i in range(len(test)):
            #     total_review += int(test[i].attrs['data-count'])
            # print(total_review, int(np.ceil(total_review/5)))

            # Article Boxes
            article_lenth = len(soup.select('article.sdp-review__article__list')) # 리뷰 한 페이지 당 5개씩

            for idx in range(article_lenth):
                # print(f"start {idx}")
                dict_data : Dict[str,Union[str,int]] = dict()
                articles = soup.select('article.sdp-review__article__list')

                # 리뷰 내용
                # 내용이 없는 리뷰는 count하지 않기 때문에 크롤링 하지 않고 넘어가야함
                review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
                if review_content == None:
                    # review_content = '등록된 리뷰내용이 없습니다'
                    continue
                else:
                    review_content = re.sub('[\n\t]','',review_content.text.strip())

                # 평점 별 클로링
                # rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                # if int(rating.attrs['data-rating']) == 1:
                #     if self.star1 >= 50:
                #         continue
                #     elif self.star1 < 50:
                #         rating, user_name, prod_name, headline, help_cnt = self.sub(articles, idx, 1)
                #         self.star1 += 1
                # elif int(rating.attrs['data-rating']) == 3:
                #     if self.star3 >= 50:
                #         continue
                #     elif self.star3 < 50:
                #         rating, user_name, prod_name, headline, help_cnt = self.sub(articles, idx, 3)
                #         self.star3 += 1
                # elif int(rating.attrs['data-rating']) == 5:
                #     if self.star5 >= 50:
                #         continue
                #     elif self.star5 < 50:
                #         rating, user_name, prod_name, headline, help_cnt = self.sub(articles, idx, 5)
                #         self.star5 += 1
                # else:
                #     continue

                rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                # print("(3) ", rating)
                if int(rating.attrs['data-rating']) == 1:
                    if self.star1 >= 50:
                        continue
                    elif self.star1 < 50:
                        rating, user_name, prod_name, headline, help_cnt = self.sub(articles, idx, 1)
                        self.star1 += 1
                elif int(rating.attrs['data-rating']) == 3:
                    if self.star3 >= 50:
                        continue
                    elif self.star3 < 50:
                        rating, user_name, prod_name, headline, help_cnt = self.sub(articles, idx, 3)
                        self.star3 += 1
                elif int(rating.attrs['data-rating']) == 5:
                    if self.star5 >= 50:
                        continue
                    elif self.star5 < 50:
                        rating, user_name, prod_name, headline, help_cnt = self.sub(articles, idx, 5)
                        self.star5 += 1
                else:
                    continue


                """
                if int(rating.attrs['data-rating']) == 1 or 3 or 5:
                    rating = int(rating.attrs['data-rating'])
                    if rating == 1:
                        star1_cnt += 1
                    elif rating == 3:
                        star3_cnt += 1
                    else:
                        star5_cnt += 1
                elif rating == None:
                    rating = 0
                else:
                    continue # 평점이 2, 4이면 해당 리뷰 크롤링 하지 않고 넘어가야함
                    # rating = int(rating.attrs['data-rating'])

                # 구매자 이름
                user_name = articles[idx].select_one('span.sdp-review__article__list__info__user__name')
                if user_name == None or user_name.text == '':
                    user_name = '-'
                else:
                    user_name = user_name.text.strip()

                # 평점
                # rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
                # if rating == None:
                #     rating = 0
                # else :
                #     rating = int(rating.attrs['data-rating'])

                # 구매자 상품명
                prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
                if prod_name == None or prod_name.text == '':
                    prod_name = '-'
                else:
                    prod_name = prod_name.text.strip()

                # 헤드라인(타이틀)
                headline = articles[idx].select_one('div.sdp-review__article__list__headline')
                if headline == None or headline.text == '':
                    headline = '등록된 헤드라인이 없습니다'
                else:
                    headline = headline.text.strip()

                # 리뷰 내용
                # review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
                # if review_content == None :
                #     review_content = '등록된 리뷰내용이 없습니다'
                # else:
                #     review_content = re.sub('[\n\t]','',review_content.text.strip())

                # 도움횟수
                help_cnt = articles[idx].select_one('div.sdp-review__article__list__help.js_reviewArticleHelpfulContainer')
                if help_cnt == None:
                    help_cnt = 0
                else:
                    help_cnt = int(help_cnt.attrs['data-count'])

                # 맛 만족도
                # answer = articles[idx].select_one('span.sdp-review__article__list__survey__row__answer')
                # if answer == None or answer.text == '':
                #     answer = '맛 평가 없음'
                # else:
                #     answer = answer.text.strip()
                """

                dict_data['prod_name'] = prod_name
                dict_data['user_name'] = user_name
                dict_data['rating'] = rating
                dict_data['headline'] = headline
                dict_data['review_content'] = review_content
                dict_data['help_cnt'] = help_cnt
                # dict_data['answer'] = answer

                save_data.append(dict_data)
                # print("!!!!!!!!!!!!!하나 크롤링 끝!!!!!!!!")
                print()
                print()
                # print(dict_data , '\n')
                # print(":::::::::::::::::::::::::::::", self.star5)
                # self.page += 1

            time.sleep(1)

            return save_data
        
    def sub(self, articles, idx, rat):
        # print()
        # if star_count >= 50:
        #     return "x", "x", "x", "x", "x"
        # else:
        rating = rat

        # 구매자 이름
        user_name = articles[idx].select_one('span.sdp-review__article__list__info__user__name')
        if user_name == None or user_name.text == '':
            user_name = '-'
        else:
            user_name = user_name.text.strip()

        # 평점
        # rating = articles[idx].select_one('div.sdp-review__article__list__info__product-info__star-orange')
        # if rating == None:
        #     rating = 0
        # else :
        #     rating = int(rating.attrs['data-rating'])

        # 구매자 상품명
        prod_name = articles[idx].select_one('div.sdp-review__article__list__info__product-info__name')
        if prod_name == None or prod_name.text == '':
            prod_name = '-'
        else:
            prod_name = prod_name.text.strip()

        # 헤드라인(타이틀)
        headline = articles[idx].select_one('div.sdp-review__article__list__headline')
        if headline == None or headline.text == '':
            headline = '등록된 헤드라인이 없습니다'
        else:
            headline = headline.text.strip()

        # 리뷰 내용
        # review_content = articles[idx].select_one('div.sdp-review__article__list__review > div')
        # if review_content == None :
        #     review_content = '등록된 리뷰내용이 없습니다'
        # else:
        #     review_content = re.sub('[\n\t]','',review_content.text.strip())

        # 도움횟수
        help_cnt = articles[idx].select_one('div.sdp-review__article__list__help.js_reviewArticleHelpfulContainer')
        if help_cnt == None:
            help_cnt = 0
        else:
            help_cnt = int(help_cnt.attrs['data-count'])

        return rating, user_name, prod_name, headline, help_cnt

    def input_review_url(self)-> str:
        while True:
            # Window
            os.system('cls')
            # Mac
            #os.system('clear')
            
            # Review URL
            review_url : str = input('원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/6451503812?itemId=14007944553&vendorItemId=73528488680&sourceType=srp_product_ads&clickEventId=28aaab30-71e3-4f30-9059-07a29eb1b27f&korePlacement=15&koreSubPlacement=6&q=%EB%9E%A9%EB%85%B8%EC%89%AC&itemsCount=36&searchId=af6bda06076947a39f847ed86a718c34&rank=5&isAddedCart=\n\n:')
            if not review_url :
                # Window
                os.system('cls')
                # Mac
                #os.system('clear')
                print('URL 주소가 입력되지 않았습니다')
                continue
            return review_url

    def input_page_count(self)-> int:
        # Window
        os.system('cls')
        # Mac
        #os.system('clear')
        while True:
            page_count : str = input('페이지 수를 입력하세요\n\n:')
            if not page_count:
                print('페이지 수가 입력되지 않았습니다\n')
                continue

            return int(page_count)

class OpenPyXL:
    @staticmethod
    def save_file()-> None:
        # 크롤링 결과
        results : List[List[Dict[str,Union[str,int]]]] = Coupang().main()

        wb = Workbook()
        ws = wb.active
        ws.append(['상품명','구매자 이름','구매자 평점','리뷰 제목','리뷰 내용','도움횟수'])

        row = 2

        for x in results:
            for result in x :
                ws[f'A{row}'] = result['prod_name']
                ws[f'B{row}'] = result['user_name']
                ws[f'C{row}'] = result['rating']
                ws[f'D{row}'] = result['headline']
                ws[f'E{row}'] = result['review_content']
                ws[f'F{row}'] = result['help_cnt']
                # ws[f'G{row}'] = result['answer']

                row += 1

        savePath : str = os.path.abspath('쿠팡-상품리뷰-크롤링')
        fileName : str = results[0][0]['prod_name'] + '.xlsx'

        if not os.path.exists(savePath):
            os.mkdir(savePath)

        wb.save(os.path.join(savePath,fileName))
        wb.close()

        print(f'파일 저장완료!\n\n{os.path.join(savePath,fileName)}')
